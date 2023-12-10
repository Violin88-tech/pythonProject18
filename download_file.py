import csv
import os.path
import zipfile, os
from io import TextIOWrapper

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from zipfile import ZipFile


CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)

tmp = os.path.join(CURRENT_DIR, "tmp")
resources = os.path.join(CURRENT_DIR, "resources")

def create_archive():
    file_dir = os.listdir(tmp)
    archive_name = os.path.join(resources, 'zip_file.zip')

    if not os.path.exists("resources"):
        os.mkdir("resources")
    elif os.path.isfile(archive_name):
        os.remove(archive_name)

    with zipfile.ZipFile(archive_name, mode='w', compression=zipfile.ZIP_DEFLATED) as archive:
        for file in file_dir:
            archive.write(os.path.join(tmp, file), file)

    return archive_name

def test_archive():
    archive_name = create_archive()
    print(archive_name)

    with zipfile.ZipFile(archive_name) as zip_file:
        print(zip_file.namelist())


        with zip_file.open('file_pdf.pdf') as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[0].extract_text()
            assert 'pytest Documentation' in text


        with zip_file.open('file_xlsx.xlsx') as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            first_title = sheet.cell(row=1, column=1).value
            assert first_title == "Внешний идентификатор для импорта"


        with zip_file.open('file_csv.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8')))
            assert 'White' == csvreader[1][0]


















